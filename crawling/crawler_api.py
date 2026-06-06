import os
import time
import re
import json
from datetime import datetime, timedelta
from collections import defaultdict
from urllib.parse import urljoin
from typing import List, Dict, Any

from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import Workbook, load_workbook

class SmartCrawler:
	def __init__(self, urls: List[str], keywords: List[str], exclude_keywords: List[str], output_excel: str = "crawling_results.xlsx", output_json: str = "crawling_summary.json"):
		self.urls = urls
		self.keywords = list(dict.fromkeys(keywords))
		self.exclude_keywords = exclude_keywords
		self.output_excel = output_excel
		self.output_json = output_json
		
		self.existing_texts = set()
		self.recent_24h_count = 0
		self.new_found_today_count = 0
		self.new_success_count = 0
		self.new_failure_count = 0
		self.recent_24h_stats = defaultdict(int)
		self.new_found_today_stats = defaultdict(int)
		
		self.driver = None
		self.wb = None
		self.ws_data = None

	def initialize_database(self) -> None:
		file_exists = os.path.isfile(self.output_excel)
		now = datetime.now()
		yesterday = now - timedelta(hours=24)

		if file_exists:
			try:
				self.wb = load_workbook(self.output_excel)
				self.ws_data = self.wb.worksheets[0]
				
				for row in self.ws_data.iter_rows(min_row=2, values_only=True):
					if not row[0]: continue
					
					status_val = str(row[1]).strip() if row[1] else ""
					kws = str(row[3]).strip() if row[3] else "알 수 없음"
					title_val = str(row[4]).strip() if row[4] else ""
					text_val = str(row[5]).strip() if row[5] else ""

					# 🚨 영구 중복 스킵을 위한 DB 메모리 적재
					if title_val: self.existing_texts.add(title_val)
					if text_val: self.existing_texts.add(text_val)
					
					try:
						row_time = datetime.strptime(str(row[0]), "%Y-%m-%d %H:%M:%S")
						if row_time >= yesterday:
							self.recent_24h_count += 1
							if status_val == "성공":
								self.recent_24h_stats[kws] += 1
					except ValueError:
						pass
				print(f"📂 기존 DB 로드 완료. (중복 스킵용 데이터 {len(self.existing_texts)}개 확보)")
			except Exception as e:
				print(f"⚠️ 기존 DB 수집 파일을 읽는 중 오류가 발생했습니다: {e}")
				file_exists = False

		if not file_exists:
			self.wb = Workbook()
			self.ws_data = self.wb.active
			self.ws_data.title = "수집데이터"
			self.ws_data.append([
				'검색 일시', '수집 상태', '검색된 사이트 URL', '발견된 키워드', 
				'전체 제목', '발견된 텍스트', '해당 주소', '링크 특이사항', '첨부파일 유무'
			])

	def setup_driver(self) -> bool:
		chrome_options = Options()
		chrome_options.add_argument('--no-sandbox')
		chrome_options.add_argument('--disable-dev-shm-usage')
		chrome_options.add_argument('user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36')
		chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
		chrome_options.add_experimental_option('useAutomationExtension', False)
		chrome_options.add_argument("--disable-blink-features=AutomationControlled")
		
		try:
			print("⚙️ 크롬 브라우저 스텔스 모드를 준비 중입니다...")
			service = Service(ChromeDriverManager().install())
			self.driver = webdriver.Chrome(service=service, options=chrome_options)
			self.driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
				"source": """ Object.defineProperty(navigator, 'webdriver', { get: () => undefined }) """
			})
			return True
		except Exception as e:
			print(f"❌ 크롬 드라이버 실행 실패: {e}")
			return False

	def run_crawler(self) -> bool:
		self.initialize_database()
		if not self.setup_driver():
			return False

		try:
			for url in self.urls:
				print(f"\n🌐 [{url}] 브라우저 접속 중...")
				try:
					self.driver.get(url)
					print("  ⏳ 안정성을 위해 화면이 완전히 뜰 때까지 15초간 대기합니다...")
					time.sleep(15)
				except Exception as e:
					print(f"  ❌ 접속 에러 (다음 사이트로 건너뜁니다): {e}")
					continue

				for keyword in self.keywords:
					print(f"\n  🔎 [현재 검색어: {keyword}] ➡ 수집 및 정밀 분석 시작")
					
					soup = BeautifulSoup(self.driver.page_source, 'html.parser')
					for script_or_style in soup(["script", "style"]):
						script_or_style.extract()

					matched_text_nodes = soup.find_all(string=re.compile(keyword))
					targets_to_click = []
					
					for text_node in matched_text_nodes:
						block_tag = text_node.parent
						while block_tag.name not in ['a', 'td', 'li', 'p', 'div', 'tr', 'body'] and block_tag.parent:
							block_tag = block_tag.parent
							
						full_text = " ".join(block_tag.get_text(separator=' ', strip=True).split())
						
						if self.exclude_keywords and any(ekw in full_text for ekw in self.exclude_keywords):
							continue 
						
						if full_text and full_text not in self.existing_texts and full_text not in [t['text'] for t in targets_to_click]:
							all_matched_kws = [kw for kw in self.keywords if kw in full_text]
							kw_combination_str = ", ".join(all_matched_kws)
							
							a_tag = block_tag if block_tag.name == 'a' else block_tag.find('a')
							if not a_tag:
								a_tag = block_tag.find_parent('a')
							
							near_url = url # 🚨 추적 불가 시 사용할 기본 메인 주소 (Fallback)
							link_attr = "없음"
							
							if a_tag:
								href = a_tag.get('href', '').strip()
								onclick = a_tag.get('onclick', '').strip()
								
								if href and not href.startswith(('#', 'javascript:')):
									near_url = urljoin(url, href) 
									link_attr = "정상링크"
								elif onclick:
									link_attr = f"이벤트: {onclick}"
								elif href:
									link_attr = f"가짜주소: {href}"
							
							targets_to_click.append({
								'text': full_text,
								'keyword_combination': kw_combination_str,
								'near_url': near_url,
								'link_attr': link_attr
							})

					if not targets_to_click:
						print(f"    📢 '{keyword}' 관련 새로운 글이 없거나 이미 DB에 수집된 상태입니다.")
						continue
						
					print(f"    🎯 새로운 글 {len(targets_to_click)}개 발견! 상세 주소 추출을 시작합니다.")

					for target in targets_to_click:
						main_window = self.driver.current_window_handle
						
						if self.driver.current_window_handle != main_window:
							self.driver.switch_to.window(main_window)
						
						safe_words = target['text'].split()[:3]
						safe_search_text = " ".join(safe_words).replace("'", "").replace('"', '').strip()
						
						if len(safe_search_text) < 2:
							safe_search_text = keyword

						current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
						kw_comb = target['keyword_combination']

						try:
							xpath = f"//*[contains(normalize-space(.), '{safe_search_text}')]"
							elements = self.driver.find_elements(By.XPATH, xpath)
							
							elem_to_click = None
							for elem in reversed(elements):
								if elem.is_displayed():
									elem_to_click = elem
									break
									
							if elem_to_click:
								before_windows = self.driver.window_handles
								time.sleep(5) # 클릭 전 버퍼
								
								self.driver.execute_script("""
									var elem = arguments[0];
									var curr = elem;
									while (curr && curr.tagName !== 'BODY') {
										if (curr.tagName === 'A' || curr.tagName === 'TR' || curr.tagName === 'BUTTON' || curr.getAttribute('onclick')) {
											if (curr.tagName === 'A') { curr.setAttribute('target', '_blank'); }
											curr.click();
											return true;
										}
										curr = curr.parentElement;
									}
									elem.click();
									return false;
								""", elem_to_click)
								
								time.sleep(5) 
								after_windows = self.driver.window_handles
								
								if len(after_windows) > len(before_windows):
									new_window = [w for w in after_windows if w not in before_windows][0]
									self.driver.switch_to.window(new_window)
									
									print(f"      ⏳ [새 창 열림] 로딩 대기 (15초)...")
									time.sleep(15)
									
									specific_url = self.driver.current_url 
									if specific_url and specific_url != "about:blank":
										self._process_success_page(specific_url, target, current_time, url, kw_comb)
									else:
										self._process_fail_page(target, current_time, url, kw_comb, "새 창 로딩 실패")
									
									self.driver.close()
									self.driver.switch_to.window(main_window)
									
								else:
									print(f"      ⏳ [현재 창 이동] 로딩 대기 (15초)...")
									time.sleep(15)
									
									specific_url = self.driver.current_url
									eval_url = specific_url.split('?')[0]
									is_main_page = any(eval_url.endswith(p) for p in ['/wholeBizMain', '/main', '/subview.do', '/list.do', '/home.do', '/selectSIIA200View.do'])
									
									if specific_url != url and not is_main_page:
										self._process_success_page(specific_url, target, current_time, url, kw_comb)
									else:
										self._process_fail_page(target, current_time, url, kw_comb, "현재 창 이동 실패 (대체 주소 보존)")
									
									self.driver.get(url)
									time.sleep(15)
							else:
								self._process_fail_page(target, current_time, url, kw_comb, "클릭 요소 추적 실패 (대체 주소 보존)")
						except Exception:
							self._process_fail_page(target, current_time, url, kw_comb, "예외 발생 (대체 주소 보존)")
							try:
								if self.driver.current_window_handle != main_window: self.driver.close()
								self.driver.switch_to.window(main_window)
							except Exception: pass
							continue
		finally:
			if self.driver:
				self.driver.quit()
			self.save_all_data()

		return True

	def _process_success_page(self, specific_url: str, target: Dict[str, Any], current_time: str, origin_url: str, kw_comb: str):
		detail_soup = BeautifulSoup(self.driver.page_source, 'html.parser')
		has_file = "없음"
		file_patterns = ['.hwp', '.hwpx', '.pdf', '.doc', '.docx', '.xls', '.xlsx', '.ppt', '.pptx', '.zip', 'download', 'downfile', 'fn_down']
		for a_tag in detail_soup.find_all('a', href=True):
			href_str = a_tag['href'].lower()
			if any(pat in href_str for pat in file_patterns):
				has_file = "있음"
				break

		self.existing_texts.add(target['text'])
		self.new_found_today_count += 1
		self.recent_24h_count += 1
		self.new_success_count += 1
		self.new_found_today_stats[kw_comb] += 1
		self.recent_24h_stats[kw_comb] += 1
		
		# 🚨 원형 100% 보존
		self.ws_data.append([
			current_time, "성공", origin_url, kw_comb, 
			target['text'], target['text'], specific_url, target['link_attr'], has_file
		])
		print(f"      ✅ [성공] 상세 주소 획득 완료! (문서파일: {has_file})")

	def _process_fail_page(self, target: Dict[str, Any], current_time: str, origin_url: str, kw_comb: str, error_msg: str):
		self.new_found_today_count += 1
		self.new_failure_count += 1
		
		# 🚨 실패 시 탭 완벽 분리 및 근접 주소(또는 대체 목록 주소) 저장
		self.ws_data.append([
			current_time, "실패", origin_url, kw_comb, 
			target['text'], target['text'], target['near_url'], target['link_attr'], "확인불가"
		])
		print(f"      ❌ [실패] {error_msg}")

	def save_all_data(self) -> None:
		def get_formatted_parts(stats_dict):
			individual, combination = [], []
			for k, v in stats_dict.items():
				if "," in k: combination.append(f"[{k}] {v}건")
				else: individual.append(f"[{k}] {v}건")
			return ", ".join(individual) if individual else "없음", ", ".join(combination) if combination else "없음"

		if "요약정보" not in self.wb.sheetnames:
			ws_summary = self.wb.create_sheet(title="요약정보")
			ws_summary.append([
				'업데이트 일시', '최근 24시간 전체건수', '신규 발견 전체건수', 
				'오늘 수집 성공건수', '오늘 수집 실패건수', 
				'24시간 단독카운팅', '24시간 조합카운팅', '신규 단독카운팅', '신규 조합카운팅'
			])
		else:
			ws_summary = self.wb["요약정보"]
		
		current_update_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
		rec_ind, rec_comb = get_formatted_parts(self.recent_24h_stats)
		new_ind, new_comb = get_formatted_parts(self.new_found_today_stats)
		
		ws_summary.append([
			current_update_time, self.recent_24h_count, self.new_found_today_count, 
			self.new_success_count, self.new_failure_count, 
			rec_ind, rec_comb, new_ind, new_comb
		])
		
		try: 
			self.wb.save(self.output_excel)
		except PermissionError: 
			print(f"\n❌ [저장 실패] DB(엑셀) 파일 접근 권한이 잠겨있습니다. 창을 닫아주세요.")

		def split_stats_for_json(stats_dict):
			individual, combination = {}, {}
			for k, v in stats_dict.items():
				if "," in k: combination[k] = v
				else: individual[k] = v
			return individual, combination

		rec_ind_json, rec_comb_json = split_stats_for_json(self.recent_24h_stats)
		new_ind_json, new_comb_json = split_stats_for_json(self.new_found_today_stats)

		summary_data = {
			"last_update": current_update_time,
			"recent_24h": {"total_count": self.recent_24h_count, "keyword_counts": rec_ind_json, "combination_counts": rec_comb_json},
			"new_found_today": {
				"total_count": self.new_found_today_count,
				"success_count": self.new_success_count, 
				"failure_count": self.new_failure_count, 
				"keyword_counts": new_ind_json, 
				"combination_counts": new_comb_json
			}
		}
		try:
			with open(self.output_json, "w", encoding="utf-8") as json_file:
				json.dump(summary_data, json_file, ensure_ascii=False, indent=4)
			print("📝 JSON 요약 파일(crawling_summary.json) DB 저장이 완료되었습니다.")
		except Exception: pass

		print("\n" + "="*50)
		print("📊 [크롤링 결과 세부 요약]")
		print(f"🕒 1. 최근 24시간 수집 전체 카운팅: {self.recent_24h_count}건")
		print("   2. 최근 24시간 키워드별 단독 카운팅:")
		for kw, count in self.recent_24h_stats.items():
			if "," not in kw: print(f"      👉 [{kw}]: {count}건")
		print("   3. 최근 24시간 키워드+키워드 조합 카운팅:")
		has_comb_24h = False
		for kw, count in self.recent_24h_stats.items():
			if "," in kw: print(f"      👉 [{kw}]: {count}건"); has_comb_24h = True
		if not has_comb_24h: print("      👉 조합 내역 없음")
			
		print(f"\n✨ 1. 이번 실행 신규 발견 전체 카운팅: {self.new_found_today_count}건")
		print(f"      ✅ [수집 성공]: {self.new_success_count}건")
		print(f"      ❌ [수집 실패]: {self.new_failure_count}건 (DB에서 순수 주소 확인 가능)")
		print("="*50)